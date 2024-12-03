from django.shortcuts import render, redirect
import django_filters
import django_tables2 as tables
from django.core.exceptions import ValidationError
from zipfile import BadZipFile
import openpyxl
from django.views.decorators.csrf import csrf_protect
from django.db import transaction
from openpyxl.utils.exceptions import InvalidFileException
import logging

from .models import Contracts, EZZC
from .forms import ImportForm
from general.common_context import common_context
import re
from general.models import Variable

logger = logging.getLogger("avantic")


def czy_same_cyfry_lub_ZT(ciagx):
    """Check if string contains only digits, 'Z' or 'T'."""
    return bool(re.match(r"^[\dZT]+$", ciagx.strip()))


def czy_same_cyfry(ciagx):
    """Check if string contains only digits."""
    return bool(re.match(r"^\d+$", ciagx.strip()))


def EZZC_row_verification(row):
    """Verifies that the row conforms to the expected format and returns the rejection reason if any."""
    if row[0] is None or row[1] is None:
        return False, "Missing critical data in columns 1 or 2."

    if isinstance(row[6], str) and row[6].strip() and not czy_same_cyfry_lub_ZT(row[6]):
        return False, "Column 7 should contain only digits, 'Z', 'T', or be empty."

    if isinstance(row[5], str) and row[5].strip() and not czy_same_cyfry(row[5]):
        return False, "Column 6 should contain only digits or be empty."

    if not isinstance(row[15], (int, float)):
        return False, "Column 16 should contain a valid number."

    if not isinstance(row[43], str) or "-" not in row[43]:
        return False, "Column 44 should contain a valid date with '-' separator."

    if not isinstance(row[44], str) or "-" not in row[44]:
        return False, "Column 45 should contain a valid date with '-' separator."

    return True, None


EXPECTED_HEADERS = [
    "Wersja",  # Version
    "Numer",  # Number
    "Umowa nadrzędna",  # Parent Contract
    "Przedmiot",  # Subject
    "Komentarze",  # Comments
    "Numer SRM",  # SRM Number
    "Numer ZZ/ZT",  # ZZ/ZT Number
    "Komentarz do działań",  # Comment on Actions
    "Opis dot. zasady przedłużenia",  # Description on Extension Rules
    "Liczba dni przed wygaśnięciem ZT",  # Number of Days Before ZZ Expiration
    "Data wpisu",  # Entry Date
    "Ilość lat objętych zamówieniem",  # Number of Years Covered by the Order
    "Odnowienie się limitu rocznego",  # Annual Limit Renewal
    "Czy wymaga ZT",  # Requires ZZ
    "Aktualizacja daty obow. Umowy",  # Update of Contract Effective Date
    "Wartość",  # Value
    "Limit roczny",  # Annual Limit
    "Wykorzystany limit roczny",  # Used Annual Limit
    "Pozostały limit roczny",  # Remaining Annual Limit
    "Dzień i miesiąc zlecenia ZT",  # ZZ Order Day and Month
    "Data podpisania",  # Signing Date
    "Data odnowienia się limitu rocznego",  # Annual Limit Renewal Date
    "Data zakończenia",  # End Date
    "Spółka",  # Company
    "Typ umowy",  # Contract Type
    "Komórka organizacyjna",  # Organizational Unit
    "Okres wypowiedzenia",  # Termination Period
    "Typ przedłużenia",  # Extension Type
    "Podstawa prawna",  # Legal Basis
    "Status",  # Status
    "Planowane/podjęte działania",  # Planned/Undertaken Actions
    "Typ wartości",  # Value Type
    "Waluta",  # Currency
    "Właściciel merytoryczny",  # Substantive Owner
    "Opiekun BZ",  # BZ Supervisor
    "Nazwa dostawcy",  # Supplier Name
    "NIP dostawcy",  # Supplier VAT Number
    "Adres dostawcy",  # Supplier Address
    "Obszary funkcjonalne",  # Functional Areas
    "Koordynatorzy",  # Coordinators
    "Inni uprawnieni",  # Other Authorized Persons
    "Typ zakresu",  # Scope Type
    "Status typu zakresu",  # Scope Type Status
    "Data obowiązywania(od kiedy)",  # Effective Date (From)
    "Data obowiązywania(do kiedy)",  # Effective Date (To)
    "Wymagany monitoring",  # Monitoring Required
    "Przypomnienie - liczba dni przed końcem daty obowiązywania",  # Reminder - Days Before Expiration
    "Wymaga powiadomienia",  # Requires Notification
    "Cykliczność powiadomienia",  # Notification Frequency
    "Czy wysyłać powiadomienie do Właściciela merytorycznego",  # Send Notification to Substantive Owner
    "Data utworzenia",  # Creation Date
]


def validate_headers(headers):
    """Validates the headers of the Excel file."""
    min_length = min(len(headers), len(EXPECTED_HEADERS))

    # Compare headers only up to the available length
    for idx in range(min_length):
        if headers[idx].strip() != EXPECTED_HEADERS[idx]:
            logger.error(
                f"Header mismatch: Expected '{EXPECTED_HEADERS[idx]}' but got '{headers[idx]}' at index {idx}"
            )
            return False

    # Check if the lengths of the headers are the same
    if len(headers) != len(EXPECTED_HEADERS):
        logger.error(
            f"Header length mismatch: Expected {len(EXPECTED_HEADERS)} headers but got {len(headers)} headers."
        )
        return False

    return True


@csrf_protect
def ezzc_import(request):
    if not request.user.groups.filter(name="contract_editor").exists():
        return redirect("/account/login")

    if request.method == "POST":
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES["excel_file"]
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active

                # Validate headers
                headers = [cell.value for cell in ws[1]]
                if not validate_headers(headers):
                    raise ValidationError("Invalid headers in the Excel file.")

                records_to_create = []
                list_display = []  # To store successfully added contracts
                rejected_list = []  # To store rejected contracts with reasons
                number_of_records = 0
                number_of_new_records = 0

                existing_sygnaturas = set(
                    EZZC.objects.values_list("sygnatura", flat=True)
                )

                with transaction.atomic():
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        sygnatura = row[1].strip() if row[1] else None
                        if sygnatura:
                            valid, reason = EZZC_row_verification(row)
                            if valid:
                                number_of_records += 1

                                if sygnatura not in existing_sygnaturas:
                                    number_of_new_records += 1
                                    new_record = EZZC(
                                        sygnatura=sygnatura,
                                        sygnatura_nadrzedna=row[2],
                                        przedmiot=row[3].strip() if row[3] else None,
                                        numer_SRM=row[5],
                                        numer_ZZZT=row[6],
                                        wartosc=row[15],
                                        typ_umowy=row[24],
                                        komorka=row[25].strip() if row[25] else None,
                                        podstawa_prawna=row[28],
                                        waluta=row[32],
                                        wlasciciel_merytoryczny=row[33],
                                        opiekun_BZ=row[34],
                                        dostawca=row[35],
                                        koordynatorzy=row[39],
                                        typ_zakresu=row[41],
                                        status=row[42].strip() if row[42] else None,
                                        od_kiedy=row[43],
                                        do_kiedy=row[44],
                                    )
                                    records_to_create.append(new_record)
                                    list_display.append(
                                        row[:13]
                                    )  # Add to accepted list
                            else:
                                rejected_list.append(
                                    (row[:13], reason)
                                )  # Add to rejected list with reason

                    if records_to_create:
                        EZZC.objects.bulk_create(records_to_create)

                wb.close()

                context = {
                    "lista": list_display,  # Accepted contracts
                    "rejected_list": rejected_list,  # Rejected contracts with reasons
                    "number_of_records": number_of_records,
                    "number_of_new_records": number_of_new_records,
                }
                context.update(common_context(request))
                Variable.set("status_importu_umow", 1)
                return render(request, "ezzc_import.html", context)

            except (
                ValidationError,
                KeyError,
                BadZipFile,
                NameError,
                TypeError,
                IndexError,
                InvalidFileException,
            ) as e:
                logger.error(f"Błąd podczas importu: {e}")
                context = {"error": str(e)}
                context.update(common_context(request))
                return render(request, "contracts/ezzc_import.html", context)

    else:
        form = ImportForm()
    context = {"form": form}
    context.update(common_context(request))
    return render(request, "contracts/import_ezzc_form.html", context)
