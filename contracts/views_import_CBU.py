from django.shortcuts import render, redirect
from django.core.exceptions import ValidationError
from zipfile import BadZipFile
from django.views.decorators.csrf import csrf_protect
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from django.db import transaction
import logging
from .models import Contracts, CBU
from .forms import ImportForm
from general.common_context import common_context
import re
from general.models import Variable

logger = logging.getLogger("avantic")

# Oczekiwane nagłówki
EXPECTED_HEADERS = [
    "Sygnatura umowy",
    "Data zawarcia umowy",
    "Data zakończenia umowy",
    "Status",
    "Nazwa kontrahenta",
    "Osoba prowadząca",
    "Wartość wydatkowa umowy netto w PLN",
    "Wartość wpływowa umowy netto w PLN",
    "Wartość odbiorów wydatkowych netto w PLN",
    "Wartość odbiorów wpływowych netto w PLN",
    "Temat",
    "iDemand",
    "Mandant",
]


def CBU_row_verification(row):
    """Walidacja wiersza dla pliku CBU z wykorzystaniem wyrażeń regularnych"""
    # Definiujemy wyrażenia regularne dla odpowiednich kolumn
    sygnatura_pattern = re.compile(r"^[A-Z]{2}/[A-Z0-9Ł ]{1,3}/\d{2}/\d{4,6}$")
    date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    end_date_pattern = re.compile(r"^(nieokreślona)$")
    status_pattern = re.compile(r"^(Zakończona|Niezakończona)$")
    errors = []
    # Sprawdzenie kolumn wg wzorców
    if not sygnatura_pattern.match(row[0]):
        errors.append("Niepoprawna sygnatura umowy")
    if not date_pattern.match(row[1]):
        errors.append("Niepoprawna data zawarcia")
    if not (date_pattern.match(row[2]) or end_date_pattern.match(row[2])):  #
        errors.append("Niepoprawna data zakończenia: " + str(row[2]))
    if not status_pattern.match(row[3]):
        errors.append("Niepoprawny status")
    # Wartości liczbowe (kolumny 6-9) muszą być większe lub równe 0
    if not all(isinstance(value, (int, float)) and value >= 0 for value in row[6:10]):
        errors.append("Jedna z wartości liczbowych jest niepoprawna")
    if errors:
        return False, errors
    return True, None


def validate_headers(headers):
    """Walidacja nagłówków pliku Excel"""
    return headers == EXPECTED_HEADERS


@csrf_protect
def cbu_import(request):
    if not request.user.groups.filter(name="contract_editor").exists():
        return redirect("/account/login")
    if request.method == "POST":
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES["excel_file"]
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active

                # Sprawdzenie nagłówków
                headers = [cell.value for cell in ws[1]]
                if not validate_headers(headers):
                    context = {"error": "Niepoprawne nagłówki pliku Excel."}
                    context.update(common_context(request))
                    return render(request, "contracts/cbu_import.html", context)
                records_to_create = []
                list_display = []
                rejected_rows = []
                number_of_records = 0
                existing_sygnatura_set = set(
                    CBU.objects.values_list("sygnatura", flat=True)
                )
                for row in ws.iter_rows(
                    min_row=2, values_only=True
                ):  # Pomijamy nagłówki
                    valid, errors = CBU_row_verification(row)
                    if valid:
                        number_of_records += 1
                        if row[0].strip() not in existing_sygnatura_set:
                            end_date = (
                                "2099-01-01" if row[2] == "nieokreślona" else row[2]
                            )
                            new_record = CBU(
                                sygnatura=row[0].strip(),
                                data_zawarcia=row[1],
                                data_zakonczenia=end_date,
                                status=row[3].strip(),
                                nazwa_kontrahenta=row[4],
                                osoba_prowadzaca=row[5],
                                wartosc_wydatkowa=row[6],
                                wartosc_wplywowa=row[7],
                                wartosc_odbiorow_wydatkowych=row[8],
                                wartosc_odbiorow_wplywowych=row[9],
                                temat=row[10].strip(),
                                idemand=row[11],
                                mandant=row[12],
                            )
                            records_to_create.append(new_record)
                            list_display.append(row[:13])
                    else:
                        rejected_rows.append((row[:13], "; ".join(errors)))
                if records_to_create:
                    with transaction.atomic():
                        CBU.objects.bulk_create(records_to_create)
                context = {
                    "lista": list_display,
                    "odrzucone": rejected_rows,
                    "number_of_records": number_of_records,
                    "number_of_new_records": len(records_to_create),
                }
                context.update(common_context(request))
                Variable.set("status_importu_umow", 1)
                return render(request, "cbu_import.html", context)

            except (ValidationError, BadZipFile, InvalidFileException, Exception) as e:
                logger.error(f"Błąd podczas importu: {e}")
                context = {"error": str(e)}
                context.update(common_context(request))
                return render(request, "contracts/cbu_import.html", context)
    else:
        form = ImportForm()
    context = {"form": form}
    context.update(common_context(request))
    return render(request, "contracts/import_cbu_form.html", context)
