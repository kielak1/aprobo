from datetime import datetime
from zipfile import BadZipFile
import re
import openpyxl

from django.shortcuts import render, redirect
from django.core.exceptions import ValidationError
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.db.models import Q
from django.views.decorators.csrf import csrf_protect

import django_tables2 as tables
from django_tables2 import SingleTableView
import django_filters

from contracts.models import Contracts, ImportedEZZ
from .models import Purchases, EZZ
from general.models import Sections, Clients
from .forms import LoginForm, UserRegistrationForm, ImportForm
from general.common_dashboard import common_dashboard
from general.common_context import common_context


@csrf_protect
def index(request):
    if not request.user.groups.filter(name="contract_viewer").exists():
        target_if_no_rights = f"/account/login"
        return redirect(target_if_no_rights)
    return common_dashboard("purchases/index.html", request)


@csrf_protect
def ezzc_add(request):
    # Pobranie rekordów z tabeli EZZC, które jeszcze nie mają powiązania w tabeli Contracts
    ezzc_without_contract = EZZC.objects.filter(ezzc__isnull=True)

    for ezzc_record in ezzc_without_contract:
        contracts_record = Contracts.objects.filter(
            cbu__sygnatura=ezzc_record.sygnatura
        ).first()
        if contracts_record:
            contracts_record.ezzc = ezzc_record
            contracts_record.save()
        else:
            contract = Contracts.objects.create(
                subject=ezzc_record.przedmiot, ezzc=ezzc_record
            )

    cbu_without_contract = CBU.objects.filter(cbu__isnull=True)
    for cbu_record in cbu_without_contract:
        contracts_record = Contracts.objects.filter(
            ezzc__sygnatura=cbu_record.sygnatura
        ).first()
        if contracts_record:
            contracts_record.cbu = cbu_record
            contracts_record.save()
        else:
            contract = Contracts.objects.create(
                subject=cbu_record.temat, cbu=cbu_record
            )

    context = {}
    return render(request, "ezzc_add.html", context)


@csrf_protect
def pusty(request):
    return render(request, "pusty.html", {})


@csrf_protect
def register(request):
    if request.method == "POST":
        user_form = UserRegistrationForm(request.POST)
        if user_form.is_valid():
            # Utworzenie nowego obiektu użytkownika, ale jeszcze nie zapisujemy go w bazie danych.
            new_user = user_form.save(commit=False)
            # Ustawienie wybranego hasła.
            new_user.set_password(user_form.cleaned_data["password"])
            # Zapisanie obiektu User.
            new_user.save()
            return render(
                request, "purchases/register_done.html", {"new_user": new_user}
            )
    else:
        user_form = UserRegistrationForm()
        return render(request, "purchases/register.html", {"user_form": user_form})


@csrf_protect
def user_login(request):
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            user = authenticate(username=cd["username"], password=cd["password"])
            if user is not None:
                if user.is_active:
                    login(request, user)
                    return HttpResponse("Uwierzytelnienie zakończyło się sukcesem.")
            else:
                return HttpResponse("Konto jest zablokowane.")
        else:
            return HttpResponse("Nieprawidłowe dane uwierzytelniające.")
    else:
        form = LoginForm()
    return render(request, "purchases/login.html", {"form": form})


class EZZFilter(django_filters.FilterSet):
    class Meta:
        model = EZZ
        fields = ("EZZ_number", "ordering_person", "subject")


class EZZTable(tables.Table):
    class Meta:
        model = EZZ
        template_name = "django_tables2/bootstrap4.html"

        attrs = {
            "class": "table"
        }  # Dodaj klasę "table" do tabeli, aby stylować ją zgodnie z Bootstrap
        per_page = 200

    ordering_person = tables.Column(
        verbose_name="Osoba zamawiająca!",
        attrs={"th": {"input": {"type": "text"}}},
    )

    subject = tables.Column(verbose_name="Temat")


class EZZListView(LoginRequiredMixin, SingleTableView):
    model = EZZ
    table_class = EZZTable
    template_name = "ezz.html"  # Nazwa twojego szablonu
    filter_class = None

    formhelper_class = None
    context_filter_name = "filter"
    login_url = "/account/login/"  # URL do przekierowania, jeśli nie zalogowany

    def dispatch(self, request, *args, **kwargs):
        if not request.user.groups.filter(name="purchase_viewer").exists():
            return redirect(self.login_url)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()

        ezz_number = self.request.GET.get("EZZ_number")
        ordering_person = self.request.GET.get("ordering_person")
        subject = self.request.GET.get("subject")

        if ezz_number:
            queryset = queryset.filter(EZZ_number=ezz_number)
        if ordering_person:
            queryset = queryset.filter(ordering_person=ordering_person)
        if subject:
            queryset = queryset.filter(subject=subject)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(common_context(self.request))
        return context


class EZZListIllegalView(LoginRequiredMixin, SingleTableView):
    model = EZZ
    table_class = EZZTable
    template_name = "ezz.html"  # Nazwa twojego szablonu
    filter_class = None

    formhelper_class = None
    context_filter_name = "filter"
    login_url = "/account/login/"  # URL do przekierowania, jeśli nie zalogowany

    def dispatch(self, request, *args, **kwargs):
        if not request.user.groups.filter(name="purchase_viewer").exists():
            return redirect(self.login_url)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()

        ezz_number = self.request.GET.get("EZZ_number")
        ordering_person = self.request.GET.get("ordering_person")
        subject = self.request.GET.get("subject")

        if ezz_number:
            queryset = queryset.filter(EZZ_number=ezz_number)
        if ordering_person:
            queryset = queryset.filter(ordering_person=ordering_person)
        if subject:
            queryset = queryset.filter(subject=subject)

        # Wykluczenie rekordów o określonych statusach w polu 'subject'
        excluded_statuses = ["Anulowany", "Roboczy", "Cofnięty do Zlecającego"]
        queryset = queryset.exclude(status__in=excluded_statuses)

        queryset = queryset.exclude(nieistotny=True)

        # Sprawdzenie, czy EZZ jest powiązany z zakupami
        queryset = queryset.filter(Q(purchases__isnull=True)).distinct()
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Dodaj parametry GET do kontekstu
        context.update(common_context(self.request))
        return context


# Define the allowed values for columns G and H based on the file analysis
ALLOWED_VALUES_G = ["OPEX", "OPEX, CAPEX", "CAPEX"]
ALLOWED_VALUES_H = ["Spółki GK ORLEN", "ORLEN"]
EXPECTED_HEADERS = [
    "Numer wniosku",
    "Zlecający",
    "Data utworzenia",
    "Przedmiot zlecenia",
    "Status",
    "Dostawca",
    "Źródło finansowania",
    "Docelowy odbiorca",
    "Obecni Akceptujący",
    "Data ostatniej akceptacji",
]


@csrf_protect
def ezz_import(request):
    if not request.user.groups.filter(name="purchase_editor").exists():
        return redirect("/account/login")

    try:
        if request.method == "POST":
            form = ImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active

                # Verify the header
                actual_headers = [
                    cell.value.strip() if cell.value else "" for cell in ws[1]
                ]  # Strip leading/trailing spaces
                expected_headers_stripped = [
                    header.strip() for header in EXPECTED_HEADERS
                ]  # Strip spaces from expected headers

                if actual_headers != expected_headers_stripped:
                    raise ValidationError("Invalid file format. Headers do not match.")

                imported_records = []
                rejected_records = []
                modified_records = []
                ImportedEZZ.objects.all().delete()

                number_of_records = 0
                number_of_new_records = 0
                number_of_modified_records = 0  # Counter for modified records

                for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
                    if not row[0]:  # Skip rows without EZZ_number
                        continue

                    number_of_records += 1

                    # Validation
                    errors = []

                    # Column A: Check if it starts with Z or T and is followed by 6 digits
                    if row[0]:
                        pattern = r"^[ZT]{2}\d{6}$"
                        if not re.match(pattern, row[0]):
                            errors.append(
                                f"EZZ_number must start with 'Z' or 'T' followed by 6 digits: {row[0]}"
                            )

                    # Column B: Uppercase, only letters, spaces
                    if row[1] and (
                        row[1] != row[1].upper()
                        or not all(c.isalpha() or c.isspace() for c in row[1])
                    ):
                        errors.append(
                            f"Current acceptor must be uppercase and contain only letters and spaces: {row[1]}"
                        )

                    # Column C: Date
                    if row[2]:
                        try:
                            datetime.strptime(str(row[2]), "%Y-%m-%d")
                        except ValueError:
                            errors.append(
                                f"Invalid date format in Creation_date: {row[2]}"
                            )

                    # # Column G: Allowed values
                    # if row[6] and row[6].strip() not in ALLOWED_VALUES_G:
                    #     errors.append(f"Invalid value in Source_of_financing: {row[6]}")

                    # Column H: Allowed values
                    # if row[7] and row[7].strip() not in ALLOWED_VALUES_H:
                    #     errors.append(f"Invalid value in Final_receiver: {row[7]}")

                    # Column I: Uppercase, only letters, spaces, and commas
                    if row[8] and (
                        row[8] != row[8].upper()
                        or not all(
                            c.isalpha() or c.isspace() or c == "," for c in row[8]
                        )
                    ):
                        errors.append(
                            f"Ordering person must be uppercase and contain only letters, spaces, and commas: {row[8]}"
                        )

                    # Column J: Date or empty
                    if row[9]:
                        try:
                            datetime.strptime(str(row[9]).strip(), "%Y-%m-%d")
                        except ValueError:
                            errors.append(
                                f"Invalid date format or non-empty value in Date_of_last_acceptance: {row[9]}"
                            )

                    if errors:
                        rejected_records.append((row, errors))
                        continue

                    # If no validation errors, process the record
                    imported_ezz = ImportedEZZ(
                        EZZ_number=row[0],
                        ordering_person=row[1],
                        creation_date=row[2],
                        subject=row[3],
                        status=row[4],
                        suplier=row[5] if isinstance(row[5], str) else None,
                        source_of_financing=(
                            row[6].strip() if isinstance(row[6], str) else None
                        ),
                        final_receiver=(
                            row[7].strip() if isinstance(row[7], str) else None
                        ),
                        current_acceptor=row[8] if isinstance(row[8], str) else None,
                        date_of_last_acceptance=(
                            row[9].strip() if isinstance(row[9], str) else None
                        ),
                    )
                    imported_ezz.save()

                    # Check if the record already exists in the EZZ model
                    try:
                        ezz_instance = EZZ.objects.get(
                            EZZ_number=imported_ezz.EZZ_number
                        )
                        existing_fields = {
                            "ordering_person": (
                                ezz_instance.ordering_person.strip()
                                if ezz_instance.ordering_person
                                else None
                            ),
                            "creation_date": (
                                ezz_instance.creation_date.strftime("%Y-%m-%d")
                                if ezz_instance.creation_date
                                else None
                            ),
                            "subject": (
                                ezz_instance.subject.strip()
                                if ezz_instance.subject
                                else None
                            ),
                            "status": (
                                ezz_instance.status.strip()
                                if ezz_instance.status
                                else None
                            ),
                            "suplier": (
                                ezz_instance.suplier.strip()
                                if ezz_instance.suplier
                                else None
                            ),
                            "source_of_financing": (
                                ezz_instance.source_of_financing.strip()
                                if ezz_instance.source_of_financing
                                else None
                            ),
                            "final_receiver": (
                                ezz_instance.final_receiver.strip()
                                if ezz_instance.final_receiver
                                else None
                            ),
                            "current_acceptor": (
                                ezz_instance.current_acceptor.strip()
                                if ezz_instance.current_acceptor
                                else None
                            ),
                            "date_of_last_acceptance": (
                                ezz_instance.date_of_last_acceptance.strftime(
                                    "%Y-%m-%d"
                                )
                                if ezz_instance.date_of_last_acceptance
                                else None
                            ),
                        }

                        imported_fields = {
                            "ordering_person": (
                                imported_ezz.ordering_person.strip()
                                if imported_ezz.ordering_person
                                else None
                            ),
                            "creation_date": (
                                str(imported_ezz.creation_date)
                                if imported_ezz.creation_date
                                else None
                            ),
                            "subject": (
                                imported_ezz.subject.strip()
                                if imported_ezz.subject
                                else None
                            ),
                            "status": (
                                imported_ezz.status.strip()
                                if imported_ezz.status
                                else None
                            ),
                            "suplier": (
                                imported_ezz.suplier.strip()
                                if imported_ezz.suplier
                                else None
                            ),
                            "source_of_financing": (
                                imported_ezz.source_of_financing.strip()
                                if imported_ezz.source_of_financing
                                else None
                            ),
                            "final_receiver": (
                                imported_ezz.final_receiver.strip()
                                if imported_ezz.final_receiver
                                else None
                            ),
                            "current_acceptor": (
                                imported_ezz.current_acceptor.strip()
                                if imported_ezz.current_acceptor
                                else None
                            ),
                            "date_of_last_acceptance": (
                                str(imported_ezz.date_of_last_acceptance).strip()
                                if imported_ezz.date_of_last_acceptance
                                else None
                            ),
                        }

                        # Check for modifications
                        if existing_fields != imported_fields:
                            # Record has been modified
                            number_of_modified_records += 1
                            modified_records.append(
                                (
                                    imported_ezz.EZZ_number,
                                    imported_ezz.ordering_person,
                                    imported_ezz.creation_date,
                                    imported_ezz.subject,
                                    imported_ezz.status,
                                    imported_ezz.suplier,
                                    imported_ezz.source_of_financing,
                                    imported_ezz.final_receiver,
                                    imported_ezz.current_acceptor,
                                    imported_ezz.date_of_last_acceptance,
                                )
                            )

                        # Update the record in the database
                        ezz_instance.ordering_person = imported_ezz.ordering_person
                        ezz_instance.creation_date = imported_ezz.creation_date
                        ezz_instance.subject = imported_ezz.subject
                        ezz_instance.status = imported_ezz.status
                        ezz_instance.suplier = imported_ezz.suplier
                        ezz_instance.source_of_financing = (
                            imported_ezz.source_of_financing
                        )
                        ezz_instance.final_receiver = imported_ezz.final_receiver
                        ezz_instance.current_acceptor = imported_ezz.current_acceptor
                        ezz_instance.date_of_last_acceptance = (
                            imported_ezz.date_of_last_acceptance
                        )
                        ezz_instance.save()

                    except EZZ.DoesNotExist:
                        # If the record does not exist, create a new one
                        ezz_instance = EZZ.objects.create(
                            EZZ_number=imported_ezz.EZZ_number,
                            ordering_person=imported_ezz.ordering_person,
                            creation_date=imported_ezz.creation_date,
                            subject=imported_ezz.subject,
                            status=imported_ezz.status,
                            suplier=imported_ezz.suplier,
                            source_of_financing=imported_ezz.source_of_financing,
                            final_receiver=imported_ezz.final_receiver,
                            current_acceptor=imported_ezz.current_acceptor,
                            date_of_last_acceptance=imported_ezz.date_of_last_acceptance,
                        )
                        number_of_new_records += 1
                        imported_records.append(
                            (
                                imported_ezz.EZZ_number,
                                imported_ezz.ordering_person,
                                imported_ezz.creation_date,
                                imported_ezz.subject,
                                imported_ezz.status,
                                imported_ezz.suplier,
                                imported_ezz.source_of_financing,
                                imported_ezz.final_receiver,
                                imported_ezz.current_acceptor,
                                imported_ezz.date_of_last_acceptance,
                            )
                        )

                wb.close()

                context = {
                    "imported_records": imported_records,
                    "rejected_records": rejected_records,
                    "modified_records": modified_records,
                    "number_of_records": number_of_records,
                    "number_of_new_records": number_of_new_records,
                    "number_of_modified_records": number_of_modified_records,
                }
                context.update(common_context(request))
                return render(request, "ezz_import.html", context)

        else:
            form = ImportForm()

        context = {"form": form}
        context.update(common_context(request))
        return render(request, "purchases/import_ezz_form.html", context)

    except (ValidationError, KeyError, BadZipFile, NameError) as e:
        print(f"Błąd: {e}")
        context = {"error": e}
        context.update(common_context(request))
        return render(request, "purchases/ezz_import.html", context)
