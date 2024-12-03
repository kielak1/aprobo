from django.shortcuts import render, redirect
from .models import Postepowania
from .forms import ImportForm
import openpyxl
from datetime import datetime
from django.core.exceptions import ValidationError
from zipfile import BadZipFile
from django.views.decorators.csrf import csrf_protect
from general.common_context import common_context
import re


# Define regex patterns for columns A, B, and I
pattern_A = re.compile(r"^\d{4,}$")  # Column A: Expecting 4 or more digits.
pattern_B = re.compile(r"^[ZT\d]+$")  # Column B: Allow digits, Z, and T.
pattern_I = re.compile(
    r"^\d{4}-\d{2}-\d{2}$"
)  # Column I: Expected date format YYYY-MM-DD.


# Function to verify the header row matches the expected format
def verify_header_row(header):
    expected_header = [
        "Numer SRM/SAP CP",
        "Numer ZZ/ZT",
        "Opis zapotrzebowania",
        "Priorytet",
        "Status SRM/SAP CP",
        "Zlecający",
        "Kupiec Biura Zakupów",
        "Status Biura Zakupów",
        "Data wprowadzenia do SRM/SAP CP",
        "Data realizacji uzgodniona przez Strony",
        "Postępowanie na Connect",
    ]
    errors = []
    for i, (expected, actual) in enumerate(zip(expected_header, header)):
        if expected != actual:
            errors.append(
                f"Kolumna {i+1}: oczekiwano '{expected}', znaleziono '{actual}'."
            )
    return not errors, errors


# New verification function for rows based on regex patterns and header check
def postepowania_row_verification(row, header_verified=False):
    if not header_verified:
        return False

    # Verify each column using regex
    col_A_valid = pattern_A.match(str(row[0]).strip())
    col_B_valid = pattern_B.match(str(row[1]).strip())

    # Verify if column I is a valid date
    col_I_valid = isinstance(row[8], datetime)

    return col_A_valid and col_B_valid and col_I_valid


@csrf_protect
def import_postepowan(request):
    if not request.user.groups.filter(name="purchase_editor").exists():
        return redirect("/account/login")

    if request.method == "POST":
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES["excel_file"]
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active

                # Verify header row
                header_row = next(ws.iter_rows(values_only=True))
                header_verified, header_errors = verify_header_row(header_row)

                if not header_verified:
                    context = {
                        "header_errors": header_errors,
                    }
                    context.update(common_context(request))
                    return render(request, "purchases/import_postepowan.html", context)

                lista = []
                rejected_rows = []
                number_of_records = 0
                number_of_new_records = 0

                # Skip the header row and process the rest
                for row in ws.iter_rows(values_only=True, min_row=2):
                    if postepowania_row_verification(row, header_verified):
                        number_of_records += 1
                        record, created = Postepowania.objects.get_or_create(
                            numer_SRM_SAP=row[0].strip(),
                            defaults={
                                "numer_ZZ": row[1].strip(),
                                "opis_zapotrzebowania": row[2],
                                "priorytet": row[3],
                                "status_SRM": row[4],
                                "zlecajacy": row[5],
                                "kupiec": row[6],
                                "status_biura": row[7],
                                "data_wprowadzenia": row[8],
                                "connect": row[10].strip() if row[10] else None,
                            },
                        )
                        if created:
                            number_of_new_records += 1
                            lista.append(row)
                        else:
                            # Update existing record
                            record.numer_ZZ = row[1].strip()
                            record.opis_zapotrzebowania = row[2]
                            record.priorytet = row[3]
                            record.status_SRM = row[4]
                            record.zlecajacy = row[5]
                            record.kupiec = row[6]
                            record.status_biura = row[7]
                            record.data_wprowadzenia = row[8]
                            record.connect = row[10].strip() if row[10] else None
                            record.save()
                    else:
                        rejected_rows.append(row)

                wb.close()

                context = {
                    "lista": lista,
                    "rejected_rows": rejected_rows,
                    "number_of_records": number_of_records,
                    "number_of_new_records": number_of_new_records,
                }
                context.update(common_context(request))
                return render(request, "purchases/import_postepowan.html", context)

            except (ValidationError, KeyError, BadZipFile, NameError, TypeError) as e:
                context = {"error": str(e)}
                context.update(common_context(request))
                return render(request, "purchases/import_postepowan.html", context)
    else:
        form = ImportForm()
    context = {"form": form}
    context.update(common_context(request))
    return render(request, "purchases/import_postepowan_form.html", context)
