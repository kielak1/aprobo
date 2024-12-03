# Django Imports
from django.db.models import Q
from django.http import HttpResponse
from django.contrib.contenttypes.models import ContentType

# Model Imports
from purchases.models import Purchases, Postepowania
from general.models import Status_procesu, Planowane_zakupy, Note
from needs.models import Needs
from ideas.models import Ideas

# OpenPyXL Imports
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_DATE_YYYYMMDD2
import io


def najblizsze_zakupy_det():
    status_purchases_zakonczony = Status_procesu.objects.get(status="zakończony")
    status_purchases_anulowany = Status_procesu.objects.get(status="anulowany")
    status_purchases_w_zakupach = Status_procesu.objects.get(status="w zakupach")
    status_purchases_zakup_BGNIG = Status_procesu.objects.get(status="zakup BGNIG")
    status_purchases_w_realizacji = Status_procesu.objects.get(status="w realizacji")

    statusy_zakupow_wykluczone = [
        status_purchases_zakonczony,
        status_purchases_anulowany,
        status_purchases_w_zakupach,
        status_purchases_zakup_BGNIG,
        status_purchases_w_realizacji,
    ]

    zakupy = Purchases.objects.filter(~Q(status_procesu__in=statusy_zakupow_wykluczone))

    print("najbliższe zakupy")
    Planowane_zakupy.objects.all().delete()

    for zakup in zakupy:

        Planowane_zakupy.objects.create(
            subject=zakup.przedmiot_zakupu,
            EZZ_number=zakup.ezz.EZZ_number,
            zakup=zakup,
            potrzeba=zakup.need,
            pomysl=Ideas.objects.filter(needs=zakup.need).first(),
            budzet=zakup.budzet_capex_netto + zakup.budzet_opex_netto,
            section=zakup.section,
            osoba_prowadzaca=zakup.osoba_prowadzaca,
            wymagana_data_realizacji=(
                zakup.need.wymagana_data_realizacji if zakup.need else None
            ),
            waluta=zakup.waluta,
            budzet_capex_netto=zakup.budzet_capex_netto,
            budzet_opex_netto=zakup.budzet_opex_netto,
            sposob_wyceny=zakup.sposob_wyceny,
            dostawca=zakup.dostawca,
            sposob_zakupu=zakup.sposob_zakupu,
            odtworzeniowy=zakup.need.odtworzeniowy if zakup.need else None,
            rozwojowy=zakup.need.rozwojowy if zakup.need else None,
            rodzaj_inicjatywy=zakup.need.rodzaj_inicjatywy if zakup.need else None,
            wymagana_data_zawarcia_umowy=zakup.wymagana_data_zawarcia_umowy,
        )

    potrzeby = Needs.objects.filter(
        Q(status_potrzeby__status="realizowana")
        | Q(status_potrzeby__status="rada architektury")
    ).exclude(purchases__isnull=False)

    for potrzeba in potrzeby:
        if potrzeba.capex:
            cp = potrzeba.capex
        else:
            cp = 0
        if potrzeba.opex:
            op = potrzeba.opex
        else:
            op = 0
        Planowane_zakupy.objects.create(
            subject=potrzeba.subject,
            zakup=None,
            potrzeba=potrzeba,
            pomysl=Ideas.objects.filter(needs=potrzeba).first(),
            budzet=cp + op,
            section=potrzeba.section,
            osoba_prowadzaca=potrzeba.osoba_prowadzaca,
            wymagana_data_realizacji=potrzeba.wymagana_data_realizacji,
            waluta=potrzeba.waluta,
            budzet_capex_netto=potrzeba.capex,
            budzet_opex_netto=potrzeba.opex,
            sposob_wyceny=potrzeba.sposob_okreslenia_budzetu,
            sposob_zakupu=potrzeba.proponowany_sposob_realizacji,
            odtworzeniowy=potrzeba.odtworzeniowy,
            rozwojowy=potrzeba.rozwojowy,
            rodzaj_inicjatywy=potrzeba.rodzaj_inicjatywy,
        )

    pomysly = Ideas.objects.filter(
        Q(status_idei__status="realizowana")
        | Q(status_idei__status="rada architektury")
    ).exclude(needs__isnull=False)

    for pomysl in pomysly:
        Planowane_zakupy.objects.create(
            subject=pomysl.subject,
            zakup=None,
            potrzeba=None,
            pomysl=pomysl,
            budzet=pomysl.orientacynjy_budzet,
            section=pomysl.section,
            osoba_prowadzaca=pomysl.osoba_prowadzaca,
            wymagana_data_realizacji=pomysl.wymagana_data_realizacji,
            waluta="",
        )


def export_planowane_zakupy_to_excel(request):
    # wygenerowanie raportu
    najblizsze_zakupy_det()

    # Stworzenie nowego arkusza Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planowane Zakupy"

    # Definicja stylów
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Dodanie nagłówków kolumn
    headers = [
        "Przedmiot zakupu",
        "EZZ",
        "Zakup",
        "Potrzeba",
        "Pomysł",
        "Budżet",
        "Waluta",
        "Dział",
        "Osoba Prowadząca",
        "Wymagana data Realizacji",
        "CAPEX",
        "OPEX",
        "wycena",
        "dostawca",
        "sposób zakupu",
        "odtworzeniowy",
        "rozwojowy",
        "rodzaj inicjatywy",
        "Wymagana data zawarcia umowy",
    ]

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Pobranie danych z modelu
    zakupy = Planowane_zakupy.objects.all()

    for row_num, zakup in enumerate(zakupy, 2):
        ws[f"A{row_num}"] = zakup.subject
        ws[f"B{row_num}"] = zakup.EZZ_number
        ws[f"C{row_num}"] = zakup.zakup.id if zakup.zakup else ""
        ws[f"D{row_num}"] = zakup.potrzeba.id if zakup.potrzeba else ""
        ws[f"E{row_num}"] = zakup.pomysl.id if zakup.pomysl else ""
        ws[f"F{row_num}"] = zakup.budzet
        ws[f"G{row_num}"] = zakup.waluta
        ws[f"H{row_num}"] = zakup.section.short_name if zakup.section else ""
        ws[f"I{row_num}"] = (
            zakup.osoba_prowadzaca.username if zakup.osoba_prowadzaca else ""
        )
        ws[f"J{row_num}"] = (
            zakup.wymagana_data_realizacji if zakup.wymagana_data_realizacji else ""
        )
        ws[f"K{row_num}"] = zakup.budzet_capex_netto if zakup.budzet_capex_netto else ""
        ws[f"L{row_num}"] = zakup.budzet_opex_netto if zakup.budzet_opex_netto else ""
        ws[f"M{row_num}"] = (
            zakup.sposob_wyceny.sposob_wyceny if zakup.sposob_wyceny else ""
        )
        ws[f"N{row_num}"] = zakup.dostawca if zakup.dostawca else ""
        ws[f"O{row_num}"] = (
            zakup.sposob_zakupu.sposob_zakupu if zakup.sposob_zakupu else ""
        )
        ws[f"P{row_num}"] = "tak" if zakup.odtworzeniowy else ""
        ws[f"Q{row_num}"] = "tak" if zakup.rozwojowy else ""
        ws[f"R{row_num}"] = (
            zakup.rodzaj_inicjatywy.rodzaj_inicjatywy if zakup.rodzaj_inicjatywy else ""
        )
        ws[f"S{row_num}"] = (
            zakup.wymagana_data_zawarcia_umowy
        )  # if zakup.rodzaj_inicjatywy else ""

        # Ustawienie zawijania tekstu, obramowań i wyśrodkowania w pionie oraz poziomie dla każdej komórki
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{row_num}"]
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
            cell.border = thin_border

            # Specjalne formatowanie dla kolumny F i J
            if col_num == 6 or col_num == 11 or col_num == 12:
                cell.number_format = "#,##0.00"  # ustawienie formatu waluty bez symbolu z separacją tysięcy spacją
            if col_num == 10:
                cell.number_format = FORMAT_DATE_YYYYMMDD2  # ustawienie formatu daty

    # Ustawienie szerokości kolumn
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 25
    ws.column_dimensions["K"].width = 18
    ws.column_dimensions["L"].width = 18
    ws.column_dimensions["N"].width = 25
    ws.column_dimensions["O"].width = 40
    ws.column_dimensions["P"].width = 16
    ws.column_dimensions["Q"].width = 16
    ws.column_dimensions["R"].width = 25
    ws.column_dimensions["S"].width = 25

    # Ustawienia odpowiedzi HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=planowane_zakupy.xlsx"

    # Zapisanie pliku Excel do odpowiedzi
    wb.save(response)

    return response


def export_trwajace_zakupy_to_excel(request):
    # wygenerowanie raportu
    najblizsze_zakupy_det()

    # Stworzenie nowego arkusza Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trwające zakupy"

    # Definicja stylów
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Dodanie nagłówków kolumn
    headers = [
        "Przedmiot zakupu",
        "EZZ",
        "Zakup",
        "Dział",
        "Wymagana data zawarcia umowy",
        "Osoba Prowadząca",
        "CAPEX",
        "OPEX",
        "Waluta",
        "wycena",
        "dostawca",
        "sposób zakupu",
        "kupiec",
        "numer SRM SAP",
        "data wprowadzenia",
        "ostatni status",
    ]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Dodanie filtra
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Pobranie danych z modelu
    zakupy = Purchases.objects.filter(status_procesu__status="w zakupach")
    purchase_content_type = ContentType.objects.get_for_model(Purchases)

    for row_num, zakup in enumerate(zakupy, 2):
        ws[f"A{row_num}"] = zakup.przedmiot_zakupu
        ws[f"B{row_num}"] = zakup.ezz.EZZ_number
        ws[f"C{row_num}"] = zakup.id
        ws[f"D{row_num}"] = zakup.section.short_name if zakup.section else ""
        ws[f"E{row_num}"] = zakup.wymagana_data_zawarcia_umowy
        ws[f"F{row_num}"] = (
            zakup.osoba_prowadzaca.username if zakup.osoba_prowadzaca else ""
        )
        ws[f"G{row_num}"] = zakup.budzet_capex_netto if zakup.budzet_capex_netto else ""
        ws[f"H{row_num}"] = zakup.budzet_opex_netto if zakup.budzet_opex_netto else ""
        ws[f"I{row_num}"] = zakup.waluta
        ws[f"J{row_num}"] = (
            zakup.sposob_wyceny.sposob_wyceny if zakup.sposob_wyceny else ""
        )
        ws[f"K{row_num}"] = zakup.dostawca if zakup.dostawca else ""
        ws[f"L{row_num}"] = (
            zakup.sposob_zakupu.sposob_zakupu if zakup.sposob_zakupu else ""
        )

        postepowanie = Postepowania.objects.filter(
            numer_ZZ=zakup.ezz.EZZ_number
        ).first()
        if postepowanie:
            ws[f"M{row_num}"] = postepowanie.kupiec
            ws[f"N{row_num}"] = postepowanie.numer_SRM_SAP
            ws[f"O{row_num}"] = postepowanie.data_wprowadzenia

        latest_note = (
            Note.objects.filter(content_type=purchase_content_type, object_id=zakup.id)
            .order_by("-timestamp")
            .first()
        )
        if latest_note:
            date_only = latest_note.timestamp.date()
            ws[f"P{row_num}"] = f"{date_only} : {latest_note.content}"

        # Ustawienie zawijania tekstu, obramowań i wyśrodkowania w pionie oraz poziomie dla każdej komórki
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{row_num}"]
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
            cell.border = thin_border

            # Specjalne formatowanie dla kolumny G i H
            if col_num == 7 or col_num == 8:
                cell.number_format = "#,##0.00"  # ustawienie formatu waluty bez symbolu z separacją tysięcy spacją
            if col_num == 5 or col_num == 15:
                cell.number_format = FORMAT_DATE_YYYYMMDD2  # ustawienie formatu daty

    # Ustawienie szerokości kolumn
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["K"].width = 20
    ws.column_dimensions["L"].width = 40
    ws.column_dimensions["M"].width = 20
    ws.column_dimensions["N"].width = 15
    ws.column_dimensions["O"].width = 20
    ws.column_dimensions["P"].width = 30

    # Zapisanie pliku Excel do wirtualnej pamięci
    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    # Ustawienia odpowiedzi HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=trwajace_zakupy.xlsx"
    response.write(virtual_workbook.read())

    return response
