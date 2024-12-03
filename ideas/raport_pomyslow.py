# Django Imports
from django.db.models import Q
from django.http import HttpResponse
from django.contrib.contenttypes.models import ContentType

# Model Imports
from ideas.models import Ideas
from general.models import Note, Clients, Sections
from needs.models import Needs
from purchases.models import Purchases, Postepowania

# OpenPyXL Imports
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_DATE_YYYYMMDD2
import io
import logging
from general.common_context import common_context


from general.linki import generate_need_url, generate_purchase_url, generate_idea_url


from openpyxl import Workbook
from openpyxl.styles import Font

logger = logging.getLogger("avantic")

def set_clickable_link(ws, cell, link, display_text=None):
    """
    Ustawia klikalny link w komórce arkusza.

    :param ws: Worksheet, arkusz Excel
    :param cell: Str, adres komórki (np. "T2")
    :param link: Str, URL
    :param display_text: Str, tekst wyświetlany zamiast URL
    """
    if link:
        ws[cell].value = display_text or link
        ws[cell].hyperlink = link
        ws[cell].font = Font(
            color="0000FF", underline="single"
        )  # Styl klikalnego linku
    else:
        ws[cell].value = ""


def export_pomysly_to_excel(request):
    # wygenerowanie raportu
    context = common_context(request)
    # Stworzenie nowego arkusza Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "inicjatywy"

    # Definicja stylów
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Dodanie nagłówków kolumn
    headers = [
        "Id",
        "Status pomysłu",
        "Status akceptacji",
        "Przedmiot inicjatywy",
        "Data utworzenia",
        "Data realizacji",
        "Orientacyjny budżet",
        "Dział",
        "Klient",
        "Osoba prowadząca",
        "Inicjator",
        "ostatnia notatka (pomysł)",
        "Need Id",
        "Status potrzeby",
        "Akceptacja potrzeby",
        "CAPEX",
        "OPEX",
        "Osoba prowadząca (potrzeba)",
        "Sposób realizacji",
        "link do Clarity",
        "link do dokumentacji",
        "CRIP ID",
        "ostatnia notatka (potrzeba)",
        "Zakup ID",
        "EZZ ID",
        "link do EZZ",
        "Status zakupu",
        "Status EZZ",
        "Bieżący akceptujący EZZ",
        "Dostawca",
        "Kupiec w Biurze Zakupów",
        "SAP CP",
        "Data wprowadzenia",
        "Connect",
        "ostatnia notatka (zakup)",
    ]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Dodanie filtra
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Pobranie danych z modelu
    #  pomysly = Ideas.objects.all()  # filter(status_procesu__status="w pomyslach")
    pomysly = Ideas.objects.exclude(
        status_idei__status__in=["zamknięta", "zrealizowana"]
    )
    logged_in_user = request.user
    if context["is_client"]:
        assigned_clients = Clients.objects.filter(users=logged_in_user)
        pomysly = pomysly.filter(client__in=assigned_clients)
    else:
        wszystkie_pomysly = context["wszystkie_pomysly"]
        if wszystkie_pomysly != "1":
            assigned_sections = Sections.objects.filter(users=logged_in_user)
            pomysly = pomysly.filter(section__in=assigned_sections)

    idea_content_type = ContentType.objects.get_for_model(Ideas)
    need_content_type = ContentType.objects.get_for_model(Needs)
    purchase_content_type = ContentType.objects.get_for_model(Purchases)

    row_num = 1
    for pomysl in pomysly:
        row_num += 1
     #   ws[f"A{row_num}"] = pomysl.id
      #  ws[f"A{row_num}"] = generate_idea_url(pomysl.id)
 
        set_clickable_link(ws, f"A{row_num}" , generate_idea_url(pomysl.id), f"{pomysl.id}")

        ws[f"B{row_num}"] = pomysl.status_idei.status if pomysl.status_idei else ""
        ws[f"C{row_num}"] = (
            pomysl.status_akceptacji.akceptacja if pomysl.status_akceptacji else ""
        )
        ws[f"D{row_num}"] = pomysl.subject
        ws[f"E{row_num}"] = pomysl.data_utworzenia
        ws[f"F{row_num}"] = pomysl.wymagana_data_realizacji
        ws[f"G{row_num}"] = pomysl.orientacynjy_budzet
        ws[f"H{row_num}"] = pomysl.section.short_name if pomysl.section else ""
        ws[f"I{row_num}"] = pomysl.client.short_name if pomysl.client else ""
        ws[f"J{row_num}"] = (
            f"{pomysl.osoba_prowadzaca.first_name} {pomysl.osoba_prowadzaca.last_name}"
            if pomysl.osoba_prowadzaca
            else ""
        )
        ws[f"K{row_num}"] = (
            f"{pomysl.inicjator.first_name} {pomysl.inicjator.last_name}"
            if pomysl.inicjator
            else ""
        )

        latest_note_idea = (
            Note.objects.filter(content_type=idea_content_type, object_id=pomysl.id)
            .order_by("-timestamp")
            .first()
        )
        if latest_note_idea:
            date_only = latest_note_idea.timestamp.date()
            ws[f"L{row_num}"] = f"{date_only} : {latest_note_idea.content}"

        connected_needs = pomysl.needs.all()
        if connected_needs.exists():
            row_num -= 1
            # Wyświetlenie powiązanych rekordów (przykładowo w logach)
            for need in connected_needs:

                row_num += 1
             #   ws[f"M{row_num}"] = need.id
                set_clickable_link(ws, f"M{row_num}" , generate_need_url(need.id), f"{need.id}")

                ws[f"N{row_num}"] = (
                    need.status_potrzeby.status if need.status_potrzeby else ""
                )
                ws[f"O{row_num}"] = (
                    need.status_akceptacji.akceptacja if need.status_akceptacji else ""
                )
                ws[f"P{row_num}"] = need.capex
                ws[f"Q{row_num}"] = need.opex
                ws[f"R{row_num}"] = (
                    f"{need.osoba_prowadzaca.first_name} {need.osoba_prowadzaca.last_name}"
                    if need.osoba_prowadzaca
                    else ""
                )
                ws[f"S{row_num}"] = (
                    need.proponowany_sposob_realizacji.sposob_zakupu
                    if need.proponowany_sposob_realizacji
                    else ""
                )
                set_clickable_link(ws, f"T{row_num}", need.link_do_clarity, "Clarity")
                set_clickable_link(
                    ws, f"U{row_num}", need.link_do_dokumentacji, "Dokumentacja"
                )

                powiazane_crip = need.pozycje_z_planu_CRIP.all()

                # Budowanie ciągu tekstowego z nazw projektów powiązanych Crip
                crip_text = (
                    "\n".join([crip.crip_id for crip in powiazane_crip])
                    if powiazane_crip
                    else ""
                )

                ws[f"V{row_num}"] = crip_text
                latest_note_need = (
                    Note.objects.filter(
                        content_type=need_content_type, object_id=need.id
                    )
                    .order_by("-timestamp")
                    .first()
                )
                if latest_note_need:
                    date_only = latest_note_need.timestamp.date()
                    ws[f"W{row_num}"] = f"{date_only} : {latest_note_need.content}"
                # Pobranie powiązanych rekordów Purchases
                connected_purchases = Purchases.objects.filter(need=need)

                if connected_purchases.exists():
                    row_num -= 1
                    # Iteracja po powiązanych zakupach i dodanie ich do arkusza Excel
                    for purchase in connected_purchases:
                        row_num += 1

                       # ws[f"X{row_num}"] = purchase.id
                        set_clickable_link(ws, f"X{row_num}" , generate_purchase_url(purchase.id), f"{purchase.id}")
                        ws[f"Y{row_num}"] = (
                            purchase.ezz.EZZ_number if purchase.ezz else ""
                        )
                        set_clickable_link(
                            ws, f"Z{row_num}", purchase.link_do_ezz, "EZZ"
                        )

                        ws[f"AA{row_num}"] = (
                            purchase.status_procesu.status
                            if purchase.status_procesu
                            else ""
                        )
                        ws[f"AB{row_num}"] = purchase.ezz.status if purchase.ezz else ""
                        ws[f"AC{row_num}"] = (
                            purchase.ezz.current_acceptor if purchase.ezz else ""
                        )
                        ws[f"AD{row_num}"] = (
                            purchase.ezz.suplier if purchase.ezz else ""
                        )
                        if purchase.ezz:
                            # Pobranie obiektu Postepowania na podstawie numeru ZZ
                            postepowanie = Postepowania.objects.filter(
                                numer_ZZ=purchase.ezz.EZZ_number
                            ).first()

                            # Sprawdzenie, czy obiekt postepowanie istnieje
                            if postepowanie:
                                ws[f"AE{row_num}"] = (
                                    postepowanie.kupiec if postepowanie.kupiec else ""
                                )
                                ws[f"AF{row_num}"] = postepowanie.numer_SRM_SAP
                                ws[f"AG{row_num}"] = postepowanie.data_wprowadzenia
                                ws[f"AH{row_num}"] = postepowanie.connect

                        latest_note_purchase = (
                            Note.objects.filter(
                                content_type=purchase_content_type,
                                object_id=purchase.id,
                            )
                            .order_by("-timestamp")
                            .first()
                        )

                        if latest_note_purchase:
                            date_only = latest_note_purchase.timestamp.date()
                            ws[f"AI{row_num}"] = (
                                f"{date_only} : {latest_note_purchase.content}"
                            )

                        # ws[f"AI{row_num}"] =  latest_note_purchase
    # Ustawienie zawijania tekstu, obramowań i wyśrodkowania w pionie oraz poziomie dla każdej komórki
    for rrnum in range(2, row_num + 1):  # Iteracja od 2 do row_num (włącznie)
        for col_num in range(1, len(headers) + 1):  # Iteracja przez wszystkie kolumny
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{rrnum}"]
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
            cell.border = thin_border

            # Specjalne formatowanie dla kolumny G
            if col_num == 7 or col_num == 15 or col_num == 16:  # Orientacyjny budżet
                cell.number_format = (
                    "#,##0.00"  # Format waluty bez symbolu z separacją tysięcy
                )
            if (
                col_num == 5 or col_num == 6
            ):  # Daty (Data utworzenia i Wymagana data realizacji)
                cell.number_format = FORMAT_DATE_YYYYMMDD2  # Format daty YYYY-MM-DD

    # Ustawienie szerokości kolumn
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 5
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 20
    ws.column_dimensions["L"].width = 40
    ws.column_dimensions["M"].width = 6
    ws.column_dimensions["N"].width = 13
    ws.column_dimensions["O"].width = 18
    ws.column_dimensions["P"].width = 12
    ws.column_dimensions["Q"].width = 12
    ws.column_dimensions["R"].width = 20
    ws.column_dimensions["S"].width = 30
    ws.column_dimensions["T"].width = 8
    ws.column_dimensions["U"].width = 15
    ws.column_dimensions["V"].width = 25
    ws.column_dimensions["W"].width = 40
    ws.column_dimensions["X"].width = 6
    ws.column_dimensions["Y"].width = 10
    ws.column_dimensions["Z"].width = 8
    ws.column_dimensions["AA"].width = 11
    ws.column_dimensions["AB"].width = 15
    ws.column_dimensions["AC"].width = 30
    ws.column_dimensions["AD"].width = 25
    ws.column_dimensions["AE"].width = 15
    ws.column_dimensions["AF"].width = 15
    ws.column_dimensions["AG"].width = 15
    ws.column_dimensions["AH"].width = 15
    ws.column_dimensions["AI"].width = 40

    # Zapisanie pliku Excel do wirtualnej pamięci
    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    # Ustawienia odpowiedzi HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=inicjatywy.xlsx"
    response.write(virtual_workbook.read())

    return response





def export_all_pomysly_to_excel(request):
    # wygenerowanie raportu
    context = common_context(request)
    # Stworzenie nowego arkusza Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "inicjatywy"

    # Definicja stylów
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Dodanie nagłówków kolumn
    headers = [
        "Id",
        "Status pomysłu",
        "Status akceptacji",
        "Przedmiot inicjatywy",
        "Data utworzenia",
        "Data realizacji",
        "Orientacyjny budżet",
        "Dział",
        "Klient",
        "Osoba prowadząca",
        "Inicjator",
        "ostatnia notatka (pomysł)",
        "Need Id",
        "Status potrzeby",
        "Akceptacja potrzeby",
        "CAPEX",
        "OPEX",
        "Osoba prowadząca (potrzeba)",
        "Sposób realizacji",
        "link do Clarity",
        "link do dokumentacji",
        "CRIP ID",
        "ostatnia notatka (potrzeba)",
        "Zakup ID",
        "EZZ ID",
        "link do EZZ",
        "Status zakupu",
        "Status EZZ",
        "Bieżący akceptujący EZZ",
        "Dostawca",
        "Kupiec w Biurze Zakupów",
        "SAP CP",
        "Data wprowadzenia",
        "Connect",
        "ostatnia notatka (zakup)",
    ]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Dodanie filtra
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Pobranie danych z modelu
    #  pomysly = Ideas.objects.all()  # filter(status_procesu__status="w pomyslach")
    pomysly = Ideas.objects.exclude(
        status_idei__status__in=["zamknięta"]
    )
    logged_in_user = request.user
    if context["is_client"]:
        assigned_clients = Clients.objects.filter(users=logged_in_user)
        pomysly = pomysly.filter(client__in=assigned_clients)
    else:
        wszystkie_pomysly = context["wszystkie_pomysly"]
        if wszystkie_pomysly != "1":
            assigned_sections = Sections.objects.filter(users=logged_in_user)
            pomysly = pomysly.filter(section__in=assigned_sections)

    idea_content_type = ContentType.objects.get_for_model(Ideas)
    need_content_type = ContentType.objects.get_for_model(Needs)
    purchase_content_type = ContentType.objects.get_for_model(Purchases)

    row_num = 1
    for pomysl in pomysly:
        row_num += 1
     #   ws[f"A{row_num}"] = pomysl.id
      #  ws[f"A{row_num}"] = generate_idea_url(pomysl.id)
 
        set_clickable_link(ws, f"A{row_num}" , generate_idea_url(pomysl.id), f"{pomysl.id}")

        ws[f"B{row_num}"] = pomysl.status_idei.status if pomysl.status_idei else ""
        ws[f"C{row_num}"] = (
            pomysl.status_akceptacji.akceptacja if pomysl.status_akceptacji else ""
        )
        ws[f"D{row_num}"] = pomysl.subject
        ws[f"E{row_num}"] = pomysl.data_utworzenia
        ws[f"F{row_num}"] = pomysl.wymagana_data_realizacji
        ws[f"G{row_num}"] = pomysl.orientacynjy_budzet
        ws[f"H{row_num}"] = pomysl.section.short_name if pomysl.section else ""
        ws[f"I{row_num}"] = pomysl.client.short_name if pomysl.client else ""
        ws[f"J{row_num}"] = (
            f"{pomysl.osoba_prowadzaca.first_name} {pomysl.osoba_prowadzaca.last_name}"
            if pomysl.osoba_prowadzaca
            else ""
        )
        ws[f"K{row_num}"] = (
            f"{pomysl.inicjator.first_name} {pomysl.inicjator.last_name}"
            if pomysl.inicjator
            else ""
        )

        latest_note_idea = (
            Note.objects.filter(content_type=idea_content_type, object_id=pomysl.id)
            .order_by("-timestamp")
            .first()
        )
        if latest_note_idea:
            date_only = latest_note_idea.timestamp.date()
            ws[f"L{row_num}"] = f"{date_only} : {latest_note_idea.content}"

        connected_needs = pomysl.needs.all()
        if connected_needs.exists():
            row_num -= 1
            # Wyświetlenie powiązanych rekordów (przykładowo w logach)
            for need in connected_needs:

                row_num += 1
             #   ws[f"M{row_num}"] = need.id
                set_clickable_link(ws, f"M{row_num}" , generate_need_url(need.id), f"{need.id}")

                ws[f"N{row_num}"] = (
                    need.status_potrzeby.status if need.status_potrzeby else ""
                )
                ws[f"O{row_num}"] = (
                    need.status_akceptacji.akceptacja if need.status_akceptacji else ""
                )
                ws[f"P{row_num}"] = need.capex
                ws[f"Q{row_num}"] = need.opex
                ws[f"R{row_num}"] = (
                    f"{need.osoba_prowadzaca.first_name} {need.osoba_prowadzaca.last_name}"
                    if need.osoba_prowadzaca
                    else ""
                )
                ws[f"S{row_num}"] = (
                    need.proponowany_sposob_realizacji.sposob_zakupu
                    if need.proponowany_sposob_realizacji
                    else ""
                )
                set_clickable_link(ws, f"T{row_num}", need.link_do_clarity, "Clarity")
                set_clickable_link(
                    ws, f"U{row_num}", need.link_do_dokumentacji, "Dokumentacja"
                )

                powiazane_crip = need.pozycje_z_planu_CRIP.all()

                # Budowanie ciągu tekstowego z nazw projektów powiązanych Crip
                crip_text = (
                    "\n".join([crip.crip_id for crip in powiazane_crip])
                    if powiazane_crip
                    else ""
                )

                ws[f"V{row_num}"] = crip_text
                latest_note_need = (
                    Note.objects.filter(
                        content_type=need_content_type, object_id=need.id
                    )
                    .order_by("-timestamp")
                    .first()
                )
                if latest_note_need:
                    date_only = latest_note_need.timestamp.date()
                    ws[f"W{row_num}"] = f"{date_only} : {latest_note_need.content}"
                # Pobranie powiązanych rekordów Purchases
                connected_purchases = Purchases.objects.filter(need=need)

                if connected_purchases.exists():
                    row_num -= 1
                    # Iteracja po powiązanych zakupach i dodanie ich do arkusza Excel
                    for purchase in connected_purchases:
                        row_num += 1

                       # ws[f"X{row_num}"] = purchase.id
                        set_clickable_link(ws, f"X{row_num}" , generate_purchase_url(purchase.id), f"{purchase.id}")
                        ws[f"Y{row_num}"] = (
                            purchase.ezz.EZZ_number if purchase.ezz else ""
                        )
                        set_clickable_link(
                            ws, f"Z{row_num}", purchase.link_do_ezz, "EZZ"
                        )

                        ws[f"AA{row_num}"] = (
                            purchase.status_procesu.status
                            if purchase.status_procesu
                            else ""
                        )
                        ws[f"AB{row_num}"] = purchase.ezz.status if purchase.ezz else ""
                        ws[f"AC{row_num}"] = (
                            purchase.ezz.current_acceptor if purchase.ezz else ""
                        )
                        ws[f"AD{row_num}"] = (
                            purchase.ezz.suplier if purchase.ezz else ""
                        )
                        if purchase.ezz:
                            # Pobranie obiektu Postepowania na podstawie numeru ZZ
                            postepowanie = Postepowania.objects.filter(
                                numer_ZZ=purchase.ezz.EZZ_number
                            ).first()

                            # Sprawdzenie, czy obiekt postepowanie istnieje
                            if postepowanie:
                                ws[f"AE{row_num}"] = (
                                    postepowanie.kupiec if postepowanie.kupiec else ""
                                )
                                ws[f"AF{row_num}"] = postepowanie.numer_SRM_SAP
                                ws[f"AG{row_num}"] = postepowanie.data_wprowadzenia
                                ws[f"AH{row_num}"] = postepowanie.connect

                        latest_note_purchase = (
                            Note.objects.filter(
                                content_type=purchase_content_type,
                                object_id=purchase.id,
                            )
                            .order_by("-timestamp")
                            .first()
                        )

                        if latest_note_purchase:
                            date_only = latest_note_purchase.timestamp.date()
                            ws[f"AI{row_num}"] = (
                                f"{date_only} : {latest_note_purchase.content}"
                            )

                        # ws[f"AI{row_num}"] =  latest_note_purchase
    # Ustawienie zawijania tekstu, obramowań i wyśrodkowania w pionie oraz poziomie dla każdej komórki
    for rrnum in range(2, row_num + 1):  # Iteracja od 2 do row_num (włącznie)
        for col_num in range(1, len(headers) + 1):  # Iteracja przez wszystkie kolumny
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}{rrnum}"]
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
            cell.border = thin_border

            # Specjalne formatowanie dla kolumny G
            if col_num == 7 or col_num == 15 or col_num == 16:  # Orientacyjny budżet
                cell.number_format = (
                    "#,##0.00"  # Format waluty bez symbolu z separacją tysięcy
                )
            if (
                col_num == 5 or col_num == 6
            ):  # Daty (Data utworzenia i Wymagana data realizacji)
                cell.number_format = FORMAT_DATE_YYYYMMDD2  # Format daty YYYY-MM-DD

    # Ustawienie szerokości kolumn
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 5
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 20
    ws.column_dimensions["L"].width = 40
    ws.column_dimensions["M"].width = 6
    ws.column_dimensions["N"].width = 13
    ws.column_dimensions["O"].width = 18
    ws.column_dimensions["P"].width = 12
    ws.column_dimensions["Q"].width = 12
    ws.column_dimensions["R"].width = 20
    ws.column_dimensions["S"].width = 30
    ws.column_dimensions["T"].width = 8
    ws.column_dimensions["U"].width = 15
    ws.column_dimensions["V"].width = 25
    ws.column_dimensions["W"].width = 40
    ws.column_dimensions["X"].width = 6
    ws.column_dimensions["Y"].width = 10
    ws.column_dimensions["Z"].width = 8
    ws.column_dimensions["AA"].width = 11
    ws.column_dimensions["AB"].width = 15
    ws.column_dimensions["AC"].width = 30
    ws.column_dimensions["AD"].width = 25
    ws.column_dimensions["AE"].width = 15
    ws.column_dimensions["AF"].width = 15
    ws.column_dimensions["AG"].width = 15
    ws.column_dimensions["AH"].width = 15
    ws.column_dimensions["AI"].width = 40

    # Zapisanie pliku Excel do wirtualnej pamięci
    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    # Ustawienia odpowiedzi HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=inicjatywy.xlsx"
    response.write(virtual_workbook.read())

    return response
